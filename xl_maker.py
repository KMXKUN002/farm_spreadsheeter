from openpyxl.workbook import Workbook
from openpyxl.styles import Font


def titler(string):
    """
    Convert programming names to human titles
    """
    string = string.title().replace('_', ' ')
    return string


def format_sheet(ws, col_names, device_data):
    """
    :param ws: Excel sheet. Each device type has one sheet.
    :param col_names: List of column names of a table
    :param device_data: List of tuples of data corresponding to the column names
    """
    # Add column names
    ws.append(titler(name) for name in col_names)

    # Add data
    for data in device_data:
        ws.append(data)

    # Bold the first row
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

    # Adjust column widths to length of values
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 1))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def create_workbook(col_names, device_data, filename):
    """
    :param col_names: Dictionary of list of column names. Device types are keys, eg. 'valves'.
    :param device_data: Dictionary of list of tuples of data. Device types are keys, eg. 'valves'.
    eg. device_data = {
        'valves': [(1, 2, 3), (2, 4, 8)],
        'ec': [(3, 6, 9), (5, 25, 125)],
        #etc.
    }

    :param filename: The raw file name, i.e., must not be a directory.
    """
    workbook = Workbook()
    index = 0
    for device_type in col_names.keys():
        ws = workbook.create_sheet(device_type, index)
        format_sheet(ws, col_names[device_type], device_data[device_type])
        index = index + 1
    workbook.save(filename)
    # print("Spreadsheet save successful")
